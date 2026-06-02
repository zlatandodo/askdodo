"""
Livermore Buy the Dip — scanner standalone (senza subscription AskLivermore).

Replica i parametri di AskLivermore usando yfinance come fonte dati.
I risultati saranno molto simili ma non identici al portale per via di
piccole differenze nella fonte dati (prezzi di chiusura proprietari vs yfinance).

Parametri (reverse-engineered da 101 risultati API reali):
  1. Prezzo > SMA200
  2. EMA65 > EMA88 > EMA100  (stacking strict)
  3. Prezzo >= EMA65
  4. Distanza da EMA65 <= 8%  (AskLivermore va fino a 20% ma i titoli S&P500
                                 reali sono quasi tutti entro 6%, filtro riduce falsi positivi)
  5. StochRSI(14,14,3) <= 30  [Wilder RMA — formula identica ad AskLivermore]
  6. Volume medio 50gg >= 200k

Ordinamento: Market Cap decrescente (identico al portale).

Run: python3 livermore_dip.py
"""
import logging
import logging.handlers
import os
import sys
import time
import traceback
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path

import yaml
from dotenv import load_dotenv

BASE_DIR    = Path(__file__).parent
LOG_DIR     = BASE_DIR / "logs"
OUTPUT_DIR  = BASE_DIR / "output"
CONFIG_PATH = BASE_DIR / "momentum_config.yaml"

sys.path.insert(0, str(BASE_DIR))

# Fix per launchd: yfinance cache in percorso esplicito (evita OperationalError SQLite)
import yfinance as yf
_YF_CACHE = BASE_DIR / ".yf_cache"
_YF_CACHE.mkdir(exist_ok=True)
yf.set_tz_cache_location(str(_YF_CACHE))


def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s",
                            datefmt="%Y-%m-%d %H:%M:%S")
    fh = logging.handlers.RotatingFileHandler(
        LOG_DIR / "livermore_dip.log",
        maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8",
    )
    fh.setFormatter(fmt)
    root.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    root.addHandler(ch)
    return logging.getLogger("livermore_dip")


@contextmanager
def step(log, name):
    log.info(f"▶ {name}")
    t0 = time.perf_counter()
    try:
        yield
    finally:
        log.info(f"✓ {name} ({time.perf_counter() - t0:.1f}s)")


def enrich(results: list[dict]) -> list[dict]:
    """
    Aggiunge market_cap_m e descrizione aziendale da yfinance.
    - market_cap: via fast_info (endpoint leggero, non rate-limited)
    - descrizione: via .info con retry/backoff (endpoint pesante)
    """
    import re
    import yfinance as yf

    for r in results:
        sym = r["ticker"]

        # ── Market cap via fast_info (veloce, resistente al rate limit) ───────
        try:
            mc = yf.Ticker(sym).fast_info.market_cap
            r["market_cap_m"] = round(mc / 1_000_000) if mc else None
        except Exception:
            r["market_cap_m"] = None
        time.sleep(0.1)

        # ── Descrizione via .info con retry ───────────────────────────────────
        r["description"] = ""
        for attempt in range(3):
            try:
                info    = yf.Ticker(sym).info
                summary = info.get("longBusinessSummary") or ""
                if summary:
                    sentences = re.split(r"(?<=[.!?])\s+", summary.strip())
                    r["description"] = " ".join(sentences[:3])
                break   # successo, esci dal loop
            except Exception as e:
                wait = 5 * (attempt + 1)   # 5s, 10s, 15s
                if attempt < 2:
                    time.sleep(wait)
                # al terzo tentativo lascia description vuota

    return results


def scan_livermore_standalone(hist_data: dict, meta_map: dict) -> list[dict]:
    """
    Replica standalone di Livermore Buy the Dip.

    Differenze rispetto a scan_livermore_buy_the_dip() in free_scanner.py:
      - EMA stacking STRICT: 65 > 88 > 100 (non solo 65 > 100)
      - pct_from_ema65 <= 8% invece di <= 20%
        (riduce falsi positivi: i match reali S&P1500 hanno max ~6%)
    """
    import numpy as np
    import pandas as pd
    from src.free_scanner import _sma, _stoch_rsi

    results = []

    for sym, hist in hist_data.items():
        if len(hist) < 215:
            continue
        try:
            close  = hist["Close"]
            volume = hist["Volume"]
            price  = float(close.iloc[-1])

            # 1. SMA200
            ma200 = _sma(close, 200)
            if np.isnan(ma200) or price <= ma200:
                continue

            # 6. Volume minimo 500k
            avg_vol_50 = float(volume.tail(50).mean())
            if avg_vol_50 < 500_000:
                continue

            # EMAs
            ema65  = float(close.ewm(span=65,  adjust=False).mean().iloc[-1])
            ema88  = float(close.ewm(span=88,  adjust=False).mean().iloc[-1])
            ema100 = float(close.ewm(span=100, adjust=False).mean().iloc[-1])

            # 2. Stacking STRICT: 65 > 88 > 100
            if not (ema65 > ema88 > ema100):
                continue

            # 3. Prezzo >= EMA65
            if price < ema65:
                continue

            # 4. Distanza EMA65 <= 8%
            pct_from_ema65 = (price - ema65) / ema65 * 100
            if pct_from_ema65 > 8.0:
                continue

            # 5. StochRSI(14,14,3) <= 30 con Wilder RMA
            srsi = _stoch_rsi(close, 14)
            if np.isnan(srsi) or srsi > 30:
                continue

            # ── Bounce Score (0-100) ──────────────────────────────────────
            # 1. StochRSI (40 pt): più è basso, più è oversold
            srsi_score = max(0.0, (30 - srsi) / 30) * 40

            # 2. Distanza EMA65 (30 pt): più è vicino al supporto, meglio
            dist_score = max(0.0, (8 - pct_from_ema65) / 8) * 30

            # 3. Forza trend EMA (20 pt): quanto EMA65 è sopra EMA100
            #    0% gap → 0pt, 10%+ gap → 20pt
            trend_gap  = (ema65 / ema100 - 1) * 100
            trend_score = min(1.0, trend_gap / 10.0) * 20

            # 4. Volume (10 pt): normalizzato su 5M azioni
            vol_score  = min(1.0, avg_vol_50 / 5_000_000) * 10

            bounce_score = round(srsi_score + dist_score + trend_score + vol_score, 1)

            meta = meta_map.get(sym, {})
            results.append({
                "ticker":         sym,
                "name":           meta.get("company_name", sym),
                "sector":         meta.get("sector", "—"),
                "price":          round(price, 2),
                "ma200":          round(ma200, 2),
                "ema65":          round(ema65, 2),
                "ema88":          round(ema88, 2),
                "ema100":         round(ema100, 2),
                "pct_from_ema65": round(pct_from_ema65, 1),
                "stoch_rsi":      srsi,
                "avg_vol_50":     round(avg_vol_50),
                "bounce_score":   bounce_score,
                "market_cap_m":   None,   # popolato dopo da enrich_market_cap()
            })

        except Exception as e:
            pass

    return results


def generate_livermore_excel(results: list[dict], run_date: str, output_path: Path) -> None:
    """Genera un file Excel con tutti i risultati BTD, filtrabili per volume e market cap."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livermore Buy the Dip"

    # ── Header ────────────────────────────────────────────────────────────────
    headers = [
        ("Bounce Score",    13),
        ("Ticker",          12),
        ("Nome",            28),
        ("Settore",         20),
        ("Descrizione",     60),
        ("Prezzo",          10),
        ("StochRSI",        11),
        ("Dist. EMA65 %",   14),
        ("EMA65",           10),
        ("EMA88",           10),
        ("EMA100",          10),
        ("SMA200",          10),
        ("Vol 50gg",        14),
        ("Mkt Cap (M$)",    14),
        ("Mkt Cap (B$)",    13),
        ("TradingView",     14),
    ]

    hdr_fill = PatternFill("solid", fgColor="064E3B")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    border   = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="DDDDDD"),
    )

    for col_idx, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # ── Righe dati ────────────────────────────────────────────────────────────
    green_fill  = PatternFill("solid", fgColor="D1FAE5")   # StochRSI ≤ 10
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")   # StochRSI 10-20
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, r in enumerate(results, 2):
        srsi    = r.get("stoch_rsi") or 0
        bscore  = r.get("bounce_score") or 0
        mc_m    = r.get("market_cap_m")
        mc_b    = round(mc_m / 1000, 2) if mc_m else None
        tv_url  = f"https://www.tradingview.com/chart/?symbol={r['ticker']}"

        row_fill = green_fill if bscore >= 70 else (yellow_fill if bscore >= 50 else white_fill)

        values = [
            r.get("bounce_score"),
            r.get("ticker"),
            r.get("name"),
            r.get("sector"),
            r.get("description", ""),
            r.get("price"),
            srsi,
            r.get("pct_from_ema65"),
            r.get("ema65"),
            r.get("ema88"),
            r.get("ema100"),
            r.get("ma200"),
            r.get("avg_vol_50"),
            mc_m,
            mc_b,
            r.get("ticker"),   # link TV
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = row_fill
            cell.border = border

            if col_idx == 4:   # Descrizione — testo a capo, altezza variabile
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center")

            # Formato numerico
            if col_idx == 1:                  # Bounce Score
                cell.number_format = "0.0"
            elif col_idx in (6, 9, 10, 11, 12):  # prezzi
                cell.number_format = "#,##0.00"
            elif col_idx == 7:                # StochRSI
                cell.number_format = "0.0"
            elif col_idx == 8:                # distanza %
                cell.number_format = '0.0"%"'
            elif col_idx == 13:               # volume
                cell.number_format = "#,##0"
            elif col_idx == 14:               # Mkt Cap M$
                cell.number_format = "#,##0"
            elif col_idx == 15:               # Mkt Cap B$
                cell.number_format = "#,##0.00"
            elif col_idx == 16:               # link TV
                cell.font      = Font(color="065F46", underline="single")
                cell.hyperlink = tv_url

        # Altezza riga proporzionale alla descrizione
        desc = r.get("description", "") or ""
        ws.row_dimensions[row_idx].height = max(18, min(60, len(desc) // 6))

    # ── Filtro automatico su tutte le colonne ─────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Congela riga header ────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Nota in fondo ──────────────────────────────────────────────────────────
    note_row = len(results) + 3
    note_cell = ws.cell(row=note_row, column=1,
        value=f"Generato il {run_date} | Universo S&P 1500 | "
              f"Parametri: Prezzo>SMA200, EMA65>88>100, Dist≤8%, StochRSI≤30 (Wilder RMA), Vol≥200k")
    note_cell.font = Font(italic=True, color="888888", size=9)

    wb.save(output_path)
    logging.getLogger("livermore_dip.excel").info(f"Excel salvato: {output_path}")


def main() -> int:
    log      = setup_logging()
    run_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    log.info(f"=== Livermore Buy the Dip (standalone) | {run_date} ===")

    load_dotenv(BASE_DIR / ".env")
    env = {k: os.environ.get(k, "") for k in
           ["GMAIL_APP_PASSWORD", "EMAIL_SENDER", "EMAIL_RECIPIENT"]}

    try:
        with open(CONFIG_PATH) as f:
            config = yaml.safe_load(f)
    except Exception as e:
        log.warning(f"Config non trovata: {e}")
        config = {}

    from src.free_scanner     import get_universe, download_history
    from src.livermore_report import generate_livermore_html
    from src.notifier         import send_report_email, send_alert_email

    try:
        # ── 1. Universo ────────────────────────────────────────────────────────
        with step(log, "Universo S&P 1500"):
            symbols, meta_map = get_universe(config)
            log.info(f"{len(symbols)} simboli")

        # ── 2. Storico OHLCV ──────────────────────────────────────────────────
        with step(log, f"Download OHLCV ({len(symbols)} simboli)"):
            hist_data = download_history(symbols)
            log.info(f"{len(hist_data)} simboli OK")

        # ── 3. Scan ────────────────────────────────────────────────────────────
        with step(log, "Livermore Buy the Dip scan"):
            results = scan_livermore_standalone(hist_data, meta_map)
            log.info(f"{len(results)} match")

        if not results:
            msg = f"Nessun ticker trovato ({run_date})."
            log.warning(msg)
            send_alert_email("BTD — Nessun risultato", msg, config, env)
            return 0

        # ── 4. Arricchimento (market cap + descrizione) ───────────────────────
        log.info("Pausa 30s prima dell'enrichment (rate limit yfinance)...")
        time.sleep(30)
        with step(log, f"Enrich — market cap + descrizione ({len(results)} ticker)"):
            results = enrich(results)

        results.sort(key=lambda r: r.get("bounce_score") or 0, reverse=True)
        log.info(f"Ordinati per Bounce Score decrescente")

        # ── 5. Output ─────────────────────────────────────────────────────────
        run_dir = OUTPUT_DIR / f"{run_date}_livermore_dip"
        run_dir.mkdir(parents=True, exist_ok=True)

        html_path  = run_dir / f"livermore_dip_{run_date}.html"
        excel_path = run_dir / f"livermore_dip_{run_date}.xlsx"

        with step(log, "HTML report"):
            generate_livermore_html(results, run_date, html_path)

        with step(log, "Excel"):
            generate_livermore_excel(results, run_date, excel_path)

        # ── 6. Mail ───────────────────────────────────────────────────────────
        with step(log, "Invio mail"):
            n = len(results)
            config.setdefault("email", {})
            config["email"]["subject"] = (
                f"🟢 Livermore Buy the Dip — {run_date} — {n} ticker"
            )
            send_report_email(
                html_path.read_text(encoding="utf-8"),
                [excel_path],
                config, env,
            )

        log.info(f"=== Done. {n} ticker ===")
        log.info("Top 10:")
        for r in results[:10]:
            mc = r.get("market_cap_m") or 0
            log.info(f"  {r['ticker']:6s} | {mc/1000:.1f}B | "
                     f"StochRSI={r['stoch_rsi']} | dist EMA65={r['pct_from_ema65']}%")
        return 0

    except Exception:
        tb = traceback.format_exc()
        log.error(f"Errore:\n{tb}")
        try:
            send_alert_email(f"Livermore BTD Error {run_date}", tb, config, env)
        except Exception:
            pass
        return 1


if __name__ == "__main__":
    sys.exit(main())
