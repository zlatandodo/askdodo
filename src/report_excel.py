"""
Generate Excel workbook with 4 sheets: Watchlist, Conviction Detail, Summary, Scanners Raw.
Uses openpyxl with formatting: navy headers, tier-colored cells, auto-sized columns.
"""
import logging
from pathlib import Path
from typing import Any

import pandas as pd

log = logging.getLogger(__name__)

# Color palette
NAVY = "1B2A4A"
WHITE = "FFFFFF"
TIER_A_GREEN = "C6EFCE"
TIER_B_YELLOW = "FFEB9C"
TIER_C_ORANGE = "FFC7CE"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW = "EFF3F8"

# Strategy colors
STRATEGY_A_BLUE = "DDEEFF"
STRATEGY_B_ORANGE = "FFE8CC"
STRATEGY_DUAL_PURPLE = "EDE0FF"


def _apply_header_style(ws, header_fill_color: str = NAVY) -> None:
    """Apply navy header style to first row of a worksheet."""
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    font = Font(color=WHITE, bold=True, size=10, name="Calibri")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _auto_size_columns(ws, min_width: int = 8, max_width: int = 40) -> None:
    """Auto-size all columns based on content."""
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _tier_fill(tier: str):
    """Return PatternFill for a tier value."""
    from openpyxl.styles import PatternFill
    color_map = {"A": TIER_A_GREEN, "B": TIER_B_YELLOW, "C": TIER_C_ORANGE}
    color = color_map.get(str(tier), "FFFFFF")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _strategy_fill(strategy: str):
    """Return PatternFill for a strategy value."""
    from openpyxl.styles import PatternFill
    color_map = {
        "A": STRATEGY_A_BLUE,
        "B": STRATEGY_B_ORANGE,
        "DUAL": STRATEGY_DUAL_PURPLE,
    }
    color = color_map.get(str(strategy), "FFFFFF")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _write_df_to_sheet(ws, df: pd.DataFrame, tier_col: str = None, strategy_col: str = None) -> None:
    """Write DataFrame to worksheet with header, optional tier coloring, and strategy coloring."""
    from openpyxl.styles import Alignment, Font, PatternFill

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=str(col_name))

    # Write data rows
    alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    plain_fill = PatternFill(fill_type=None)
    data_font = Font(size=9, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center")

    tier_col_idx = None
    if tier_col and tier_col in df.columns:
        tier_col_idx = list(df.columns).index(tier_col)

    strategy_col_idx = None
    if strategy_col and strategy_col in df.columns:
        strategy_col_idx = list(df.columns).index(strategy_col)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        row_tier = row.get(tier_col) if tier_col else None
        row_strategy = row.get(strategy_col) if strategy_col else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Format numeric values
            if isinstance(value, float) and not pd.isna(value):
                cell.value = round(value, 4)
            elif pd.isna(value) if not isinstance(value, str) else False:
                cell.value = None
            else:
                cell.value = value

            cell.font = data_font
            cell.alignment = center_align

            if col_idx - 1 == tier_col_idx and row_tier:
                cell.fill = _tier_fill(str(row_tier))
            elif col_idx - 1 == strategy_col_idx and row_strategy:
                cell.fill = _strategy_fill(str(row_strategy))
            elif row_idx % 2 == 0:
                cell.fill = alt_fill
            else:
                cell.fill = plain_fill

    _apply_header_style(ws)
    _auto_size_columns(ws)

    # Freeze top row
    ws.freeze_panes = ws["A2"]


def _build_watchlist_df(watchlist_df: pd.DataFrame) -> pd.DataFrame:
    """Select and rename columns for the main Watchlist sheet."""
    cols = [
        "symbol", "company_name", "sector", "tier", "strategy", "conviction_score",
        "price", "entry", "stop", "target1", "target2",
        "risk_per_share", "size_usd", "size_shares", "size_pct", "risk_usd",
        "rr_t1", "rr_t2", "atr14", "ma10", "ma50", "ma200",
        "ars", "ta", "fa", "market_cap_m",
        "next_earnings_date", "days_to_earnings", "scanners_hit", "filter_note",
    ]
    available = [c for c in cols if c in watchlist_df.columns]
    return watchlist_df[available].copy()


def _build_conviction_df(conviction_df: pd.DataFrame) -> pd.DataFrame:
    """Build Conviction Detail sheet with scanner presence columns."""
    base_cols = [
        "symbol", "company_name", "sector", "tier", "conviction_score",
        "primary_count", "confirm_count", "ars", "ta", "fa",
    ]
    available_base = [c for c in base_cols if c in conviction_df.columns]
    scanner_presence_cols = [
        c for c in conviction_df.columns
        if c not in base_cols and c.startswith(("high_", "power_", "vcp", "episodic_",
                                                  "livermore_", "golden_", "sector_",
                                                  "institutional_", "insider_", "bounce_"))
    ]
    return conviction_df[available_base + scanner_presence_cols].copy()


def generate_excel(
    watchlist_df: pd.DataFrame,
    conviction_df: pd.DataFrame,
    portfolio_summary: dict,
    scanner_stats: dict,
    output_path: Path,
) -> Path:
    """
    Generate multi-sheet Excel workbook with watchlist and supporting data.
    Returns path to the written file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # --- Sheet 1: Watchlist ---
    ws1 = wb.active
    ws1.title = "Watchlist"
    watch_df = _build_watchlist_df(watchlist_df)
    _write_df_to_sheet(ws1, watch_df, tier_col="tier", strategy_col="strategy")
    log.info(f"Sheet 'Watchlist' written: {len(watch_df)} rows.")

    # --- Sheet 2: Conviction Detail ---
    ws2 = wb.create_sheet("Conviction Detail")
    conv_df = _build_conviction_df(conviction_df)
    _write_df_to_sheet(ws2, conv_df, tier_col="tier")
    log.info(f"Sheet 'Conviction Detail' written: {len(conv_df)} rows.")

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet("Summary")
    _apply_header_style(ws3)

    summary_rows = [
        ["Metric", "Value"],
        ["Run Date", portfolio_summary.get("run_date", "")],
        ["Capital (USD)", portfolio_summary.get("capital_usd", "")],
        ["N Positions", portfolio_summary.get("n_positions", "")],
        ["Total Allocated (USD)", portfolio_summary.get("total_allocated_usd", "")],
        ["Exposure (%)", portfolio_summary.get("exposure_pct", "")],
        ["Total Risk (USD)", portfolio_summary.get("total_risk_usd", "")],
        ["Total Risk (%)", portfolio_summary.get("total_risk_pct", "")],
        ["Risk Per Trade (%)", portfolio_summary.get("risk_per_trade_pct", "")],
        ["Tier A Count", portfolio_summary.get("tier_a_count", "")],
        ["Tier B Count", portfolio_summary.get("tier_b_count", "")],
    ]
    for row_data in summary_rows:
        ws3.append(row_data)
    _apply_header_style(ws3)
    _auto_size_columns(ws3)

    # --- Sheet 4: Scanners Raw ---
    ws4 = wb.create_sheet("Scanners Raw")
    if scanner_stats:
        stats_rows = [["Scanner", "CSV Available", "Tickers Downloaded"]]
        for name, info in scanner_stats.items():
            stats_rows.append([name, info.get("available", False), info.get("count", 0)])
        for row_data in stats_rows:
            ws4.append(row_data)
        _apply_header_style(ws4)
        _auto_size_columns(ws4)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    log.info(f"Excel report saved: {output_path}")
    return output_path
