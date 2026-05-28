"""
HTML report + Excel for the Momentum Focus List and Breakout Alert.
Light theme — email-client compatible (Gmail, Apple Mail, Outlook).
"""
import logging
from pathlib import Path

import pandas as pd

log = logging.getLogger(__name__)

# ── CSS email-safe (light theme, no grid, no dark backgrounds) ───────────────

CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Arial, sans-serif;
  background: #f0f4f8;
  color: #1a202c;
  font-size: 14px;
}
.container { max-width: 820px; margin: 0 auto; padding: 24px 16px; }

header {
  background: linear-gradient(135deg, #1a365d 0%, #2a4a8a 100%);
  color: #fff;
  padding: 28px 32px;
  border-radius: 10px;
  margin-bottom: 20px;
}
header h1 { font-size: 22px; font-weight: 800; }
header .sub { font-size: 12px; color: #bee3f8; margin-top: 6px; }
.regime-badge {
  display: inline-block;
  font-size: 12px; font-weight: 700;
  padding: 4px 14px; border-radius: 20px; margin-top: 10px;
}
.regime-bull { background: #c6f6d5; color: #22543d; }
.regime-weak { background: #fed7d7; color: #742a2a; }

.regime-bar {
  background: #fff; border-left: 4px solid #4299e1;
  border-radius: 6px; padding: 10px 16px;
  margin-bottom: 18px; font-size: 12px; color: #4a5568; line-height: 1.8;
}

.stats-row {
  display: table; width: 100%; border-spacing: 10px; margin-bottom: 18px;
}
.stat-cell {
  display: table-cell;
  background: #fff; border-radius: 8px;
  padding: 14px 16px; text-align: center;
  box-shadow: 0 1px 3px rgba(0,0,0,0.08);
  border-top: 3px solid #4299e1;
}
.stat-cell .s-label { font-size: 10px; color: #718096; text-transform: uppercase; letter-spacing: 0.5px; }
.stat-cell .s-val   { font-size: 22px; font-weight: 800; color: #2b6cb0; margin-top: 4px; }

.links-bar {
  background: #fff; border-radius: 8px;
  padding: 12px 16px; margin-bottom: 16px;
  font-size: 12px; color: #4a5568; line-height: 2.2;
  box-shadow: 0 1px 3px rgba(0,0,0,0.07);
}
.links-bar strong { font-size: 10px; text-transform: uppercase;
  letter-spacing: 0.5px; color: #2b6cb0; display: block; margin-bottom: 3px; }
.links-bar a { color: #2b6cb0; text-decoration: none; font-weight: 700; }

.finviz-btn {
  display: inline-block; background: #2b6cb0; color: #fff;
  text-decoration: none; font-size: 12px; font-weight: 700;
  padding: 8px 18px; border-radius: 6px; margin-bottom: 14px;
}

/* Setup cards */
.card {
  background: #fff; border-radius: 10px;
  border: 1px solid #e2e8f0; border-left: 5px solid #4299e1;
  padding: 18px 20px; margin-bottom: 14px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.07);
}

.card-head { display: flex; align-items: baseline;
  gap: 10px; margin-bottom: 8px; flex-wrap: wrap; }
.card-rank { color: #a0aec0; font-size: 13px; font-weight: 800; }
.card-sym  { color: #2b6cb0; font-weight: 900; font-size: 18px; text-decoration: none; }
.card-sym:hover { text-decoration: underline; }
.card-company { color: #718096; font-size: 13px; }
.card-sector  {
  font-size: 10px; font-weight: 700; padding: 2px 8px;
  background: #ebf8ff; color: #2b6cb0; border-radius: 4px;
}

.card-desc { color: #4a5568; font-size: 12px; line-height: 1.65; margin-bottom: 12px; }

.metrics { display: flex; gap: 18px; flex-wrap: wrap; margin-bottom: 10px; }
.met .m-lbl { font-size: 10px; color: #a0aec0; text-transform: uppercase; letter-spacing: 0.4px; }
.met .m-val { font-size: 13px; font-weight: 700; color: #2d3748; margin-top: 2px; }
.met .m-val.blue   { color: #2b6cb0; }
.met .m-val.green  { color: #276749; }
.met .m-val.orange { color: #c05621; }

.tags { display: flex; gap: 6px; flex-wrap: wrap; margin-top: 6px; }
.tag {
  font-size: 10px; font-weight: 700; padding: 3px 9px; border-radius: 4px;
}
.tag-move  { background: #ebf4ff; color: #2b6cb0; }
.tag-tight { background: #f0fff4; color: #276749; }
.tag-vol   { background: #fef3c7; color: #92400e; }

.score-badge {
  float: right; text-align: center;
  background: #ebf4ff; border-radius: 8px; padding: 8px 14px; margin-left: 12px;
}
.score-badge .s-num { font-size: 26px; font-weight: 900; color: #2b6cb0; }
.score-badge .s-lbl { font-size: 9px; color: #a0aec0; text-transform: uppercase; letter-spacing: 0.5px; }
.score-bar { height: 4px; border-radius: 2px;
  background: linear-gradient(90deg, #4299e1, #2b6cb0); margin-top: 4px; }

.ema-mini { font-size: 10px; color: #718096; margin-top: 6px; line-height: 1.7; }

/* Breakout */
.bo-card {
  background: #fff; border-radius: 10px;
  border: 2px solid #38a169; padding: 20px;
  margin-bottom: 14px; box-shadow: 0 1px 3px rgba(0,0,0,0.07);
}
.bo-head { display: flex; align-items: center; gap: 10px; flex-wrap: wrap; margin-bottom: 8px; }
.bo-sym  { color: #276749; font-weight: 900; font-size: 20px; text-decoration: none; }
.bo-badge { background: #c6f6d5; color: #22543d; font-size: 11px;
  font-weight: 700; padding: 3px 10px; border-radius: 12px; }

footer { text-align: center; color: #a0aec0; font-size: 11px; margin-top: 28px; }
"""


def _fmt(val, decimals=2, suffix="", prefix="") -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "—"
    try:
        return f"{prefix}{float(val):,.{decimals}f}{suffix}"
    except Exception:
        return str(val)


def _fmt_vol(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "—"
    try:
        v = float(val)
        if v >= 1_000_000_000: return f"{v/1e9:.1f}B"
        if v >= 1_000_000:     return f"{v/1e6:.1f}M"
        if v >= 1_000:         return f"{v/1000:.0f}K"
        return str(int(v))
    except Exception:
        return "—"


def _fmt_mcap(val_m) -> str:
    if val_m is None or (isinstance(val_m, float) and pd.isna(val_m)):
        return "—"
    try:
        m = float(val_m)
        return f"${m/1000:.1f}B" if m >= 1000 else f"${m:.0f}M"
    except Exception:
        return "—"


def generate_momentum_html(
    focus_list: list[dict],
    regime: dict,
    run_date: str,
    output_path: Path,
) -> Path:
    n = len(focus_list)
    regime_bullish = regime.get("bullish", False)
    regime_label   = "✅ MERCATO BULLISH — full conviction" if regime_bullish else "⚠️ MERCATO DEBOLE — ridurre size"
    regime_cls     = "regime-bull" if regime_bullish else "regime-weak"

    details = regime.get("details", {})
    spy_d   = details.get("SPY", {})
    qqq_d   = details.get("QQQ", {})
    regime_detail = ""
    if spy_d:
        regime_detail += (f"<strong>SPY</strong> {spy_d.get('price','—')} &nbsp;|&nbsp; "
                          f"EMA21 {spy_d.get('ema21','—')} &nbsp;|&nbsp; EMA50 {spy_d.get('ema50','—')} &nbsp;&nbsp;&nbsp;")
    if qqq_d:
        regime_detail += (f"<strong>QQQ</strong> {qqq_d.get('price','—')} &nbsp;|&nbsp; "
                          f"EMA21 {qqq_d.get('ema21','—')} &nbsp;|&nbsp; EMA50 {qqq_d.get('ema50','—')}")

    symbols      = [c["symbol"] for c in focus_list]
    finviz_url   = "https://finviz.com/screener.ashx?v=111&t=" + ",".join(symbols) if symbols else "#"
    tv_links_html = " &nbsp;|&nbsp; ".join(
        f'<a href="https://www.tradingview.com/chart/?symbol={s}" target="_blank">{s}</a>'
        for s in symbols
    )

    avg_score = round(sum(c.get("quality_score", 0) for c in focus_list) / n, 1) if n else 0
    avg_prior = round(sum(c.get("prior_move_pct", 0) for c in focus_list) / n, 1) if n else 0
    max_score = max((c.get("quality_score", 1) for c in focus_list), default=100)

    stats_html = f"""
    <div class="stats-row">
      <div class="stat-cell"><div class="s-label">Setup trovati</div><div class="s-val">{n}</div></div>
      <div class="stat-cell"><div class="s-label">Avg Score</div><div class="s-val">{avg_score}</div></div>
      <div class="stat-cell"><div class="s-label">Avg Prior Move</div><div class="s-val">+{avg_prior}%</div></div>
      <div class="stat-cell"><div class="s-label">Regime</div><div class="s-val" style="color:{'#276749' if regime_bullish else '#c53030'}">{"BULL" if regime_bullish else "WEAK"}</div></div>
    </div>"""

    cards_html = ""
    for rank, c in enumerate(focus_list, 1):
        sym     = c["symbol"]
        tv_url  = f"https://www.tradingview.com/chart/?symbol={sym}"
        company = c.get("company_name", "")
        sector  = c.get("sector") or "—"
        score   = c.get("quality_score", 0)
        desc    = c.get("description") or "<em style='color:#a0aec0'>Descrizione non disponibile.</em>"
        price   = _fmt(c.get("price"), 2, prefix="$")
        ema8    = _fmt(c.get("ema8"),  2, prefix="$")
        ema21   = _fmt(c.get("ema21"), 2, prefix="$")
        ema50   = _fmt(c.get("ema50"), 2, prefix="$")
        prior   = _fmt(c.get("prior_move_pct"), 1, suffix="%", prefix="+")
        base_r  = _fmt(c.get("base_range_pct"), 1, suffix="%")
        vol_r   = _fmt(c.get("vol_ratio"), 2, suffix="×")
        adr     = _fmt(c.get("adr20_pct"), 1, suffix="%")
        revenue = f"${c['revenue_b']}B" if c.get("revenue_b") else "—"
        vol_wk  = _fmt_vol(c.get("vol_week"))
        mcap    = _fmt_mcap(c.get("market_cap_m"))
        bar_w   = max(6, int(score / max_score * 80))

        cards_html += f"""
        <div class="card">
          <div class="score-badge">
            <div class="s-num">{score}</div>
            <div class="s-lbl">score</div>
            <div class="score-bar" style="width:{bar_w}px"></div>
            <div class="ema-mini">EMA8 {ema8}<br>EMA21 {ema21}<br>EMA50 {ema50}</div>
          </div>
          <div class="card-head">
            <span class="card-rank">#{rank}</span>
            <a class="card-sym" href="{tv_url}" target="_blank">{sym} ↗</a>
            <span class="card-company">{company}</span>
            <span class="card-sector">{sector}</span>
          </div>
          <div class="card-desc">{desc}</div>
          <div class="metrics">
            <div class="met"><div class="m-lbl">Prezzo</div><div class="m-val blue">{price}</div></div>
            <div class="met"><div class="m-lbl">Fatturato</div><div class="m-val">{revenue}</div></div>
            <div class="met"><div class="m-lbl">Mkt Cap</div><div class="m-val">{mcap}</div></div>
            <div class="met"><div class="m-lbl">Vol sett.</div><div class="m-val orange">{vol_wk}</div></div>
            <div class="met"><div class="m-lbl">ADR</div><div class="m-val">{adr}</div></div>
          </div>
          <div class="tags">
            <span class="tag tag-move">📈 Move {prior}</span>
            <span class="tag tag-tight">🔲 Base {base_r}</span>
            <span class="tag tag-vol">📊 Vol {vol_r}</span>
          </div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Momentum Focus List — {run_date}</title>
  <style>{CSS}</style>
</head>
<body>
<div class="container">

  <header>
    <h1>🚀 Momentum Focus List</h1>
    <div class="sub">{run_date} &nbsp;·&nbsp; {n} setup &nbsp;·&nbsp; Sean's Swing Framework &nbsp;·&nbsp; S&amp;P 1500</div>
    <div class="regime-badge {regime_cls}">{regime_label}</div>
  </header>

  <div class="regime-bar"><strong>Market Regime: </strong>{regime_detail}</div>

  {stats_html}

  <a class="finviz-btn" href="{finviz_url}" target="_blank">📊 Apri tutti su Finviz</a>

  <div class="links-bar">
    <strong>📈 TradingView — apri singolo chart</strong>
    {tv_links_html}
  </div>

  {cards_html}

  <footer>
    Momentum Focus List &nbsp;·&nbsp; {run_date}<br>
    Solo a scopo informativo. Non è consulenza finanziaria.
  </footer>
</div>
</body>
</html>"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    log.info(f"Momentum HTML saved: {output_path}")
    return output_path


def generate_momentum_excel(
    focus_list: list[dict],
    run_date: str,
    output_path: Path,
) -> Path:
    """Genera Excel con tutti i dati della focus list, filtrabile."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Momentum Focus List"

    headers = [
        ("#",              5),
        ("Ticker",        10),
        ("Azienda",       28),
        ("Settore",       20),
        ("Descrizione",   55),
        ("Score",          8),
        ("Prezzo",        10),
        ("EMA8",          10),
        ("EMA21",         10),
        ("EMA50",         10),
        ("Prior Move %",  13),
        ("Base Range %",  13),
        ("Vol Ratio",     11),
        ("ADR %",          9),
        ("Fatturato (B$)", 14),
        ("Mkt Cap (B$)",  13),
        ("Vol Settimana",  14),
        ("TradingView",   12),
    ]

    hdr_fill = PatternFill("solid", fgColor="1A365D")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin     = Side(style="thin", color="DDDDDD")
    border   = Border(bottom=thin, right=thin)

    for col_idx, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    # Score colori
    fills = {
        "high":   PatternFill("solid", fgColor="C6F6D5"),  # score > 35
        "mid":    PatternFill("solid", fgColor="FEFCBF"),  # score 25-35
        "low":    PatternFill("solid", fgColor="FFF5F5"),  # score < 25
    }

    for row_idx, c in enumerate(focus_list, 2):
        score = c.get("quality_score", 0) or 0
        row_fill = fills["high"] if score > 35 else (fills["mid"] if score >= 25 else fills["low"])

        tv_url  = f"https://www.tradingview.com/chart/?symbol={c['symbol']}"
        rev     = c.get("revenue_b")
        mc_m    = c.get("market_cap_m")
        mc_b    = round(mc_m / 1000, 2) if mc_m else None

        values = [
            row_idx - 1,
            c.get("symbol"),
            c.get("company_name", ""),
            c.get("sector", ""),
            c.get("description", "") or "",
            score,
            c.get("price"),
            c.get("ema8"),
            c.get("ema21"),
            c.get("ema50"),
            c.get("prior_move_pct"),
            c.get("base_range_pct"),
            c.get("vol_ratio"),
            c.get("adr20_pct"),
            rev,
            mc_b,
            c.get("vol_week"),
            c.get("symbol"),   # TV link
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = row_fill
            cell.border = border
            cell.alignment = Alignment(
                vertical="top" if col_idx == 5 else "center",
                wrap_text=(col_idx == 5),
            )
            # Formati
            if col_idx in (7, 8, 9, 10):     cell.number_format = "#,##0.00"
            elif col_idx in (11, 12, 13, 14): cell.number_format = "0.0"
            elif col_idx == 15:               cell.number_format = "#,##0.0"   # Fatturato B$
            elif col_idx == 16:               cell.number_format = "#,##0.00"  # Mkt Cap B$
            elif col_idx == 17:               cell.number_format = "#,##0"
            elif col_idx == 18:             # TV link
                cell.font      = Font(color="2B6CB0", underline="single")
                cell.hyperlink = tv_url

        desc = c.get("description", "") or ""
        ws.row_dimensions[row_idx].height = max(20, min(70, len(desc) // 5))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

    note_row = len(focus_list) + 3
    ws.cell(row=note_row, column=1,
        value=f"Generato {run_date} | Parametri: EMA8>21>50, ADR>2%, Base<12%, Vol contrazione, Prior move>10%"
    ).font = Font(italic=True, color="888888", size=9)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(f"Momentum Excel saved: {output_path}")
    return output_path


def generate_breakout_html(
    breakouts: list[dict],
    regime: dict,
    run_date: str,
    output_path: Path,
) -> Path:
    n = len(breakouts)
    regime_bullish = regime.get("bullish", False)
    regime_cls     = "regime-bull" if regime_bullish else "regime-weak"
    regime_label   = "✅ MERCATO BULLISH" if regime_bullish else "⚠️ MERCATO DEBOLE"

    breakouts  = sorted(breakouts, key=lambda b: b.get("vol_ratio_today") or 0, reverse=True)
    symbols    = [b["symbol"] for b in breakouts]
    finviz_url = "https://finviz.com/screener.ashx?v=111&t=" + ",".join(symbols) if symbols else "#"

    cards_html = ""
    for b in breakouts:
        sym     = b["symbol"]
        company = b.get("company_name", "")
        sector  = b.get("sector") or "—"
        tv_url  = f"https://www.tradingview.com/chart/?symbol={sym}"
        desc    = b.get("description") or ""
        price   = _fmt(b.get("today_close"),      2, prefix="$")
        level   = _fmt(b.get("breakout_level"),   2, prefix="$")
        pct     = _fmt(b.get("breakout_pct"),     1, suffix="%", prefix="+")
        vol_r   = _fmt(b.get("vol_ratio_today"),  1, suffix="×")
        vol_wk  = _fmt_vol(b.get("vol_week"))
        revenue = f"${b['revenue_b']}B" if b.get("revenue_b") else "—"
        mcap    = _fmt_mcap(b.get("market_cap_m"))

        # colore volume: verde se >=2x, arancione se >=1.5x, blu se >=1.05x
        vr_raw = b.get("vol_ratio_today") or 0
        vol_color = "#276749" if vr_raw >= 2.0 else ("#c05621" if vr_raw >= 1.5 else "#2b6cb0")

        cards_html += f"""
        <div class="bo-card">
          <div class="bo-head">
            <a class="bo-sym" href="{tv_url}" target="_blank">{sym} ↗</a>
            <span style="color:#718096;font-size:13px">{company}</span>
            <span class="bo-badge">🔥 BREAKOUT {pct}</span>
            <span class="card-sector">{sector}</span>
          </div>
          <div style="display:flex;align-items:center;gap:24px;margin:8px 0 10px 0">
            <div style="text-align:center">
              <div style="font-size:11px;color:#718096;text-transform:uppercase;letter-spacing:0.5px">Volume vs media</div>
              <div style="font-size:32px;font-weight:800;color:{vol_color};line-height:1.1">{vol_r}</div>
            </div>
            <div style="text-align:center">
              <div style="font-size:11px;color:#718096;text-transform:uppercase;letter-spacing:0.5px">Breakout</div>
              <div style="font-size:32px;font-weight:800;color:#276749;line-height:1.1">{pct}</div>
            </div>
            <div style="text-align:center">
              <div style="font-size:11px;color:#718096;text-transform:uppercase;letter-spacing:0.5px">Prezzo</div>
              <div style="font-size:32px;font-weight:800;color:#1a365d;line-height:1.1">{price}</div>
            </div>
          </div>
          <div class="card-desc" style="margin-bottom:10px">{desc}</div>
          <div class="metrics">
            <div class="met"><div class="m-lbl">Livello rotto</div><div class="m-val">{level}</div></div>
            <div class="met"><div class="m-lbl">Vol ultima sett.</div><div class="m-val orange">{vol_wk}</div></div>
            <div class="met"><div class="m-lbl">Fatturato</div><div class="m-val">{revenue}</div></div>
            <div class="met"><div class="m-lbl">Mkt Cap</div><div class="m-val">{mcap}</div></div>
          </div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>⚡ Breakout Alert — {run_date}</title>
  <style>{CSS}</style>
</head>
<body>
<div class="container">
  <header>
    <h1>⚡ Breakout Alert</h1>
    <div class="sub">{run_date} &nbsp;·&nbsp; {n} breakout dalla focus list</div>
    <div class="regime-badge {regime_cls}">{regime_label}</div>
  </header>
  <a class="finviz-btn" href="{finviz_url}" target="_blank">📊 Apri su Finviz</a>
  {cards_html}
  <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 16px;font-size:11px;color:#718096;line-height:1.8;margin-top:8px">
    <strong>Criteri breakout:</strong>
    Close &gt; max High ultimi 20gg +1% &nbsp;·&nbsp;
    Volume &gt; 1.05× media base &nbsp;·&nbsp;
    Prezzo sopra EMA21 e EMA50 &nbsp;·&nbsp;
    Ordinati per volume decrescente
  </div>
  <footer>Breakout Alert &nbsp;·&nbsp; {run_date} &nbsp;·&nbsp; Non è consulenza finanziaria.</footer>
</div>
</body>
</html>"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    log.info(f"Breakout HTML saved: {output_path}")
    return output_path
