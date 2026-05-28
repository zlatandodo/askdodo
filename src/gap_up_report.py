"""
Gap Up Report — genera HTML + Excel per i Buyable Gap Up.
Aggiornato per la v3: gap recenti (lookback 10gg), senza filtri RS/base/SMA200.
"""
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


# ── HTML helpers ──────────────────────────────────────────────────────────────

def _grade_badge(quality: str) -> str:
    colors = {
        "A+": ("14532d", "86efac"),
        "A":  ("1e3a5f", "93c5fd"),
        "B+": ("7c2d12", "fdba74"),
        "B":  ("374151", "d1d5db"),
    }
    tc, bg = colors.get(quality, ("374151", "d1d5db"))
    return (f'<span style="background:#{bg};color:#{tc};font-weight:700;'
            f'padding:2px 8px;border-radius:12px;font-size:12px;">{quality}</span>')


def _rs_badge(rs_rating: int) -> str:
    if rs_rating >= 90:
        bg, tc = "14532d", "86efac"
    elif rs_rating >= 80:
        bg, tc = "1e3a5f", "93c5fd"
    elif rs_rating >= 70:
        bg, tc = "713f12", "fde68a"
    else:
        bg, tc = "374151", "d1d5db"
    return (f'<span style="background:#{bg};color:#{tc};font-weight:700;'
            f'padding:2px 8px;border-radius:12px;font-size:12px;">{rs_rating}</span>')


def _rvol_color(rvol: float) -> str:
    if rvol >= 5.0: return "#15803d"
    if rvol >= 3.0: return "#22c55e"
    if rvol >= 2.0: return "#f97316"
    return "#64748b"


def _days_badge(days_ago: int) -> str:
    if days_ago == 1: bg, tc, label = "15803d", "ffffff", "OGGI"
    elif days_ago <= 3: bg, tc, label = "1e3a5f", "ffffff", f"{days_ago}gg fa"
    elif days_ago <= 7: bg, tc, label = "713f12", "fde68a", f"{days_ago}gg fa"
    else: bg, tc, label = "374151", "d1d5db", f"{days_ago}gg fa"
    return (f'<span style="background:#{bg};color:#{tc};font-weight:600;'
            f'padding:2px 7px;border-radius:10px;font-size:11px;">{label}</span>')


def _trend_icon(above_50: bool, above_200: bool) -> str:
    if above_50 and above_200:
        return '<span style="color:#22c55e">▲▲</span>'
    if above_50:
        return '<span style="color:#f97316">▲</span>'
    return '<span style="color:#ef4444">▼</span>'


# ── HTML principale ───────────────────────────────────────────────────────────

def generate_gap_up_html(results: list[dict], run_date: str,
                         output_path: Path) -> None:
    """Genera il report HTML dei Buyable Gap Up."""

    n_ap = sum(1 for r in results if r["quality"] == "A+")
    n_a  = sum(1 for r in results if r["quality"] == "A")
    n_bp = sum(1 for r in results if r["quality"] == "B+")
    n_b  = sum(1 for r in results if r["quality"] == "B")

    stats_html = f"""
    <div class="stats-row">
      <div class="stat-card"><div class="stat-val green">{n_ap}</div><div class="stat-lbl">A+</div></div>
      <div class="stat-card"><div class="stat-val blue">{n_a}</div><div class="stat-lbl">A</div></div>
      <div class="stat-card"><div class="stat-val orange">{n_bp}</div><div class="stat-lbl">B+</div></div>
      <div class="stat-card"><div class="stat-val gray">{n_b}</div><div class="stat-lbl">B</div></div>
      <div class="stat-card"><div class="stat-val">{len(results)}</div><div class="stat-lbl">Totale BGU</div></div>
    </div>"""

    rows_html = ""
    for r in results:
        tv_url  = f"https://www.tradingview.com/chart/?symbol={r['ticker']}"
        tv_link = f'<a href="{tv_url}" target="_blank" style="color:#3b82f6;text-decoration:none;">📈</a>'

        rvol_color = _rvol_color(r["rvol"])
        trend_icon = _trend_icon(r["above_sma50"], r["above_sma200"])
        mc_m  = r["market_cap_m"]
        mc_str = f"${mc_m/1000:.1f}B" if mc_m >= 1000 else f"${mc_m:.0f}M" if mc_m > 0 else "—"

        # Gap held colore
        gh = r["gap_held_pct"]
        gh_color = "#15803d" if gh >= 80 else ("#f97316" if gh >= 40 else "#dc2626")
        gh_str = f'<span style="color:{gh_color};font-weight:600;">{gh:.0f}%</span>'

        # Variazione giornaliera
        dc = r["day_chg_pct"]
        dc_color = "#15803d" if dc > 0 else "#dc2626"
        dc_str = f'<span style="color:{dc_color};font-weight:600;">{dc:+.2f}%</span>'

        # Gap zone: zona da comprare
        gap_zone = f'${r["gap_zone_low"]:.2f} – ${r["gap_zone_high"]:.2f}'

        rows_html += f"""
        <tr>
          <td style="text-align:center;">{_grade_badge(r['quality'])}</td>
          <td style="text-align:center;font-weight:700;color:#1e293b;">{r['gap_score']}</td>
          <td style="text-align:center;">{_days_badge(r['gap_days_ago'])}</td>
          <td>
            <div style="font-weight:700;color:#0f172a;">{r['ticker']} {tv_link}</div>
            <div style="font-size:11px;color:#64748b;">{r['name'][:28]}</div>
          </td>
          <td style="font-size:11px;color:#64748b;">{r['sector']}</td>
          <td style="font-weight:700;color:#22c55e;text-align:center;">+{r['gap_pct']:.2f}%</td>
          <td style="text-align:center;">{gh_str}</td>
          <td style="text-align:center;color:{rvol_color};font-weight:700;">{r['rvol']:.2f}×</td>
          <td style="font-size:11px;color:#64748b;text-align:center;">{gap_zone}</td>
          <td style="text-align:right;font-weight:600;">${r['price']:.2f}</td>
          <td style="text-align:center;">{dc_str}</td>
          <td style="text-align:center;">{trend_icon}</td>
          <td style="text-align:center;">{_rs_badge(r['rs_rating'])}</td>
          <td style="text-align:right;font-size:11px;">{r['volume']/1e6:.2f}M</td>
          <td style="text-align:right;font-size:11px;">{mc_str}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Buyable Gap Up — {run_date}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: #f8fafc; color: #1e293b; padding: 24px; }}
  .header {{ background: linear-gradient(135deg, #0c1445 0%, #1e3a5f 100%);
             color: white; padding: 28px 32px; border-radius: 12px; margin-bottom: 20px; }}
  .header h1 {{ font-size: 24px; font-weight: 800; margin-bottom: 4px; }}
  .header p  {{ font-size: 13px; color: #93c5fd; }}
  .method-note {{ background: #f0f9ff; border-left: 4px solid #3b82f6;
                  padding: 10px 16px; border-radius: 4px; margin-bottom: 16px;
                  font-size: 12px; color: #1e3a5f; }}
  .stats-row {{ display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }}
  .stat-card {{ background: white; border-radius: 8px; padding: 14px 20px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.08); flex: 1; min-width: 80px; text-align: center; }}
  .stat-val  {{ font-size: 26px; font-weight: 800; color: #1e293b; }}
  .stat-lbl  {{ font-size: 11px; color: #94a3b8; font-weight: 600; margin-top: 2px; text-transform: uppercase; }}
  .stat-val.green  {{ color: #22c55e; }}
  .stat-val.blue   {{ color: #3b82f6; }}
  .stat-val.orange {{ color: #f97316; }}
  .stat-val.gray   {{ color: #94a3b8; }}
  .table-wrap {{ background: white; border-radius: 12px;
                 box-shadow: 0 1px 3px rgba(0,0,0,0.08); overflow: auto; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 12.5px; }}
  th {{ background: #1e3a5f; color: white; padding: 10px 10px; font-weight: 700;
        font-size: 11px; text-transform: uppercase; letter-spacing: 0.04em;
        white-space: nowrap; text-align: left; position: sticky; top: 0; }}
  td {{ padding: 9px 10px; border-bottom: 1px solid #f1f5f9; white-space: nowrap; }}
  tr:hover td {{ background: #f8fafc; }}
  tr:last-child td {{ border-bottom: none; }}
  .legend {{ margin-top: 16px; font-size: 11px; color: #64748b; line-height: 1.8; }}
</style>
</head>
<body>

<div class="header">
  <h1>📊 Buyable Gap Up Scanner</h1>
  <p>{run_date} — {len(results)} BGU trovati — Universo S&amp;P 1500 — Metodologia AskLivermore</p>
</div>

<div class="method-note">
  <b>Logica:</b>
  Gap ≥ 3% (open vs prev.close) negli ultimi 10 giorni &nbsp;·&nbsp;
  Volume sul gap day ≥ 1.5× media 20gg &nbsp;·&nbsp;
  Gap zone ancora intatta (prezzo attuale &gt; close pre-gap) &nbsp;·&nbsp;
  Ordinamento per Gap Score (freschezza + RVOL + gap size + gap held)
</div>

{stats_html}

<div class="table-wrap">
<table>
<thead>
<tr>
  <th>Quality</th>
  <th>Score</th>
  <th title="Quando è avvenuto il gap">Gap Age</th>
  <th>Ticker</th>
  <th>Settore</th>
  <th title="Gap %: open vs close del giorno precedente">Gap %</th>
  <th title="Gap Held: quanto del gap è ancora intatto (prezzo vs prev.close pre-gap)">Held %</th>
  <th title="RVOL: volume sul gap day vs media 20gg pre-gap">RVOL</th>
  <th title="Gap Zone: zona tra la close pre-gap e l'open del gap — area da comprare sul pullback">Gap Zone</th>
  <th>Prezzo</th>
  <th title="Variazione % nell'ultima sessione">Oggi %</th>
  <th title="Trend SMA50/200">Trend</th>
  <th title="RS Rating 12m — percentile 1-99 (solo informativo)">ARS</th>
  <th>Volume</th>
  <th>Mkt Cap</th>
</tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>
</div>

<div class="legend">
  <b>Gap Zone:</b> zona compresa tra close pre-gap e open del gap — buy sul pullback in questa zona &nbsp;|&nbsp;
  <b>Gap Held %:</b> (prezzo attuale - close pre-gap) / (open gap - close pre-gap) × 100 &nbsp;|&nbsp;
  <b>RVOL:</b> volume sul gap day / media 20gg pre-gap &nbsp;|&nbsp;
  <b>ARS:</b> RS Rating 12m (percentile 1-99) — solo informativo, non usato come filtro
</div>

</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")
    log.info(f"HTML salvato: {output_path}")


# ── Excel ──────────────────────────────────────────────────────────────────────

def generate_gap_up_excel(results: list[dict], run_date: str,
                          output_path: Path) -> None:
    """Genera il report Excel dei Buyable Gap Up."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"BGU {run_date}"

    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="E2E8F0")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    grade_fills = {
        "A+": PatternFill("solid", fgColor="86EFAC"),
        "A":  PatternFill("solid", fgColor="93C5FD"),
        "B+": PatternFill("solid", fgColor="FDBA74"),
        "B":  PatternFill("solid", fgColor="D1D5DB"),
    }
    grade_fonts = {
        "A+": Font(bold=True, color="14532D"),
        "A":  Font(bold=True, color="1E3A5F"),
        "B+": Font(bold=True, color="7C2D12"),
        "B":  Font(bold=True, color="374151"),
    }
    green_font = Font(color="15803D", bold=True)
    red_font   = Font(color="DC2626", bold=True)
    gray_font  = Font(color="64748B")
    alt_fill   = PatternFill("solid", fgColor="F8FAFC")

    headers = [
        "Quality", "Gap Score", "Gap Age (gg)",
        "Ticker", "Azienda", "Settore",
        "Gap %", "Gap Held %", "RVOL",
        "Gap Zone Low ($)", "Gap Zone High ($)",
        "Prezzo", "Oggi %",
        "SMA50", "SMA200", "Sopra SMA50", "Sopra SMA200",
        "ARS Rating",
        "Volume", "Vol Medio 20gg", "Mkt Cap (B$)",
        "TradingView",
    ]
    col_widths = [
        9, 10, 13,
        9, 26, 22,
        9, 11, 9,
        16, 16,
        10, 10,
        10, 10, 12, 13,
        10,
        14, 16, 13,
        20,
    ]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for ri, r in enumerate(results, 2):
        tv_url = f"https://www.tradingview.com/chart/?symbol={r['ticker']}"
        mc_b   = round(r["market_cap_m"] / 1000, 2) if r["market_cap_m"] else 0.0

        row_vals = [
            r["quality"],
            r["gap_score"],
            r["gap_days_ago"],
            r["ticker"],
            r["name"],
            r["sector"],
            r["gap_pct"] / 100,
            r["gap_held_pct"] / 100,
            r["rvol"],
            r["gap_zone_low"],
            r["gap_zone_high"],
            r["price"],
            r["day_chg_pct"] / 100,
            r["sma50"],
            r["sma200"],
            "✓" if r["above_sma50"] else "✗",
            "✓" if r["above_sma200"] else "✗",
            r["rs_rating"],
            r["volume"],
            r["avg_vol_20"],
            mc_b,
            tv_url,
        ]

        is_alt = (ri % 2 == 0)
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_alt:
                cell.fill = alt_fill

        # Quality
        ws.cell(row=ri, column=1).fill = grade_fills.get(r["quality"], PatternFill())
        ws.cell(row=ri, column=1).font = grade_fonts.get(r["quality"], Font(bold=True))

        # Gap Age colore
        d = r["gap_days_ago"]
        age_fill = (PatternFill("solid", fgColor="86EFAC") if d == 1 else
                    PatternFill("solid", fgColor="93C5FD") if d <= 3 else
                    PatternFill("solid", fgColor="FDE68A") if d <= 7 else
                    PatternFill())
        ws.cell(row=ri, column=3).fill = age_fill
        ws.cell(row=ri, column=3).font = Font(bold=True)

        # Gap %
        ws.cell(row=ri, column=7).font          = green_font
        ws.cell(row=ri, column=7).number_format = "0.00%"

        # Gap Held %
        gh = r["gap_held_pct"]
        ws.cell(row=ri, column=8).font = green_font if gh >= 80 else (red_font if gh < 40 else gray_font)
        ws.cell(row=ri, column=8).number_format = "0.0%"

        # RVOL
        rv = r["rvol"]
        ws.cell(row=ri, column=9).font = (green_font if rv >= 5.0 else
                                          Font(color="F97316", bold=True) if rv >= 2.0 else gray_font)
        ws.cell(row=ri, column=9).number_format = "0.00"

        # Gap Zone
        ws.cell(row=ri, column=10).number_format = "$#,##0.00"
        ws.cell(row=ri, column=11).number_format = "$#,##0.00"

        # Prezzi
        ws.cell(row=ri, column=12).number_format = "$#,##0.00"
        ws.cell(row=ri, column=12).font = Font(bold=True)

        # Oggi %
        dc = r["day_chg_pct"]
        ws.cell(row=ri, column=13).font = green_font if dc > 0 else red_font
        ws.cell(row=ri, column=13).number_format = "+0.00%;-0.00%"

        # SMA
        ws.cell(row=ri, column=14).number_format = "$#,##0.00"
        ws.cell(row=ri, column=15).number_format = "$#,##0.00"

        # RS Rating colore
        rs = r["rs_rating"]
        rs_fill = (PatternFill("solid", fgColor="86EFAC") if rs >= 90 else
                   PatternFill("solid", fgColor="93C5FD") if rs >= 80 else
                   PatternFill("solid", fgColor="FDE68A") if rs >= 70 else
                   PatternFill())
        ws.cell(row=ri, column=18).fill = rs_fill
        ws.cell(row=ri, column=18).font = Font(bold=True)

        # Volume
        ws.cell(row=ri, column=19).number_format = "#,##0"
        ws.cell(row=ri, column=20).number_format = "#,##0"

        # Mkt Cap
        ws.cell(row=ri, column=21).number_format = "#,##0.00"

        # TV link
        lc = ws.cell(row=ri, column=22)
        lc.hyperlink = tv_url
        lc.font = Font(color="3B82F6", underline="single")

        ws.row_dimensions[ri].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results)+1}"
    wb.save(str(output_path))
    log.info(f"Excel salvato: {output_path}")
