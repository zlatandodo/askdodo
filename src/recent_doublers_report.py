"""
Recent Doublers Report — genera HTML + Excel.
"""
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


# ── HTML helpers ──────────────────────────────────────────────────────────────

def _grade_badge(quality: str) -> str:
    colors = {
        "A+": ("14532d", "86efac"),
        "A":  ("1e3a5f", "93c5fd"),
        "B+": ("7c2d12", "fdba74"),
        "B":  ("374151", "d1d5db"),
    }
    tc, bg = colors.get(quality, ("374151", "d1d5db"))
    return (f'<span style="background:#{bg};color:#{tc};font-weight:700;'
            f'padding:2px 8px;border-radius:12px;font-size:12px;">{quality}</span>')


def _rs_badge(rs: int) -> str:
    if rs >= 90:   bg, tc = "14532d", "86efac"
    elif rs >= 80: bg, tc = "1e3a5f", "93c5fd"
    elif rs >= 70: bg, tc = "713f12", "fde68a"
    else:          bg, tc = "374151", "d1d5db"
    return (f'<span style="background:#{bg};color:#{tc};font-weight:700;'
            f'padding:2px 8px;border-radius:12px;font-size:12px;">{rs}</span>')


def _ret_cell(pct: float) -> str:
    if pct == 0:
        return '<span style="color:#94a3b8;">—</span>'
    color = "#15803d" if pct >= 200 else ("#22c55e" if pct >= 100 else "#64748b")
    bold  = "font-weight:700;" if pct >= 100 else ""
    return f'<span style="color:{color};{bold}">+{pct:.0f}%</span>'


def _dist_cell(pct: float) -> str:
    color = "#22c55e" if pct <= 5 else ("#f97316" if pct <= 15 else "#ef4444")
    return f'<span style="color:{color};font-weight:600;">-{pct:.1f}%</span>'


def _speed_badge(tf: str) -> str:
    colors = {"3m": ("065f46","86efac"), "6m": ("1e3a5f","93c5fd"),
              "9m": ("7c2d12","fdba74"), "12m": ("374151","d1d5db")}
    tc, bg = colors.get(tf, ("374151","d1d5db"))
    return (f'<span style="background:#{bg};color:#{tc};font-weight:700;'
            f'padding:2px 7px;border-radius:10px;font-size:11px;">{tf}</span>')


# ── HTML ──────────────────────────────────────────────────────────────────────

def generate_recent_doublers_html(results: list[dict], run_date: str,
                                   output_path: Path) -> None:
    n_ap = sum(1 for r in results if r["quality"] == "A+")
    n_a  = sum(1 for r in results if r["quality"] == "A")
    n_bp = sum(1 for r in results if r["quality"] == "B+")
    n_b  = sum(1 for r in results if r["quality"] == "B")
    avg_rs = round(sum(r["rs_rating"] for r in results) / len(results)) if results else 0

    stats_html = f"""
    <div class="stats-row">
      <div class="stat-card"><div class="stat-val green">{n_ap}</div><div class="stat-lbl">A+ (≤3m)</div></div>
      <div class="stat-card"><div class="stat-val blue">{n_a}</div><div class="stat-lbl">A (≤6m)</div></div>
      <div class="stat-card"><div class="stat-val orange">{n_bp}</div><div class="stat-lbl">B+ (≤9m)</div></div>
      <div class="stat-card"><div class="stat-val gray">{n_b}</div><div class="stat-lbl">B (≤12m)</div></div>
      <div class="stat-card"><div class="stat-val">{len(results)}</div><div class="stat-lbl">Totale</div></div>
      <div class="stat-card"><div class="stat-val">{avg_rs}</div><div class="stat-lbl">RS medio</div></div>
    </div>"""

    rows_html = ""
    for r in results:
        tv_url  = f"https://www.tradingview.com/chart/?symbol={r['ticker']}"
        tv_link = f'<a href="{tv_url}" target="_blank" style="color:#3b82f6;text-decoration:none;font-weight:700;">{r["ticker"]} ↗</a>'
        mc_b    = r["market_cap_m"] / 1000 if r["market_cap_m"] else 0
        mc_str  = f"${mc_b:.1f}B" if mc_b >= 1 else f"${r['market_cap_m']:.0f}M"
        sma_icon = "✅" if r["above_sma50"] else "❌"

        rows_html += f"""
        <tr>
          <td style="text-align:center;">{_grade_badge(r['quality'])}</td>
          <td style="text-align:center;font-weight:700;">{r['doubler_score']}</td>
          <td style="text-align:center;">{_speed_badge(r['fastest_tf'])}</td>
          <td style="text-align:center;">{_rs_badge(r['rs_rating'])}</td>
          <td>
            <div>{tv_link}</div>
            <div style="font-size:11px;color:#64748b;">{r['name'][:28]}</div>
          </td>
          <td style="font-size:11px;color:#64748b;">{r['sector']}</td>
          <td style="text-align:center;">{_ret_cell(r['ret_3m'])}</td>
          <td style="text-align:center;">{_ret_cell(r['ret_6m'])}</td>
          <td style="text-align:center;">{_ret_cell(r['ret_12m'])}</td>
          <td style="text-align:center;font-weight:700;color:#15803d;">+{r['best_return']:.0f}%</td>
          <td style="text-align:center;">{_dist_cell(r['dist_52w_pct'])}</td>
          <td style="text-align:right;font-weight:600;">${r['price']:.2f}</td>
          <td style="text-align:center;">{sma_icon}</td>
          <td style="text-align:right;font-size:11px;">{r['avg_vol_50']/1e6:.2f}M</td>
          <td style="text-align:right;font-size:11px;">{mc_str}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<title>Recent Doublers — {run_date}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: #f8fafc; color: #1e293b; padding: 24px; }}
  .header {{ background: linear-gradient(135deg, #065f46 0%, #047857 100%);
             color: white; padding: 28px 32px; border-radius: 12px; margin-bottom: 20px; }}
  .header h1 {{ font-size: 24px; font-weight: 800; margin-bottom: 4px; }}
  .header p  {{ font-size: 13px; color: #6ee7b7; }}
  .method-note {{ background: #f0fdf4; border-left: 4px solid #22c55e;
                  padding: 10px 16px; border-radius: 4px; margin-bottom: 16px;
                  font-size: 12px; color: #14532d; }}
  .stats-row {{ display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }}
  .stat-card {{ background: white; border-radius: 8px; padding: 14px 20px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.08); flex: 1; min-width: 80px; text-align: center; }}
  .stat-val  {{ font-size: 26px; font-weight: 800; color: #1e293b; }}
  .stat-lbl  {{ font-size: 11px; color: #94a3b8; font-weight: 600; margin-top: 2px; text-transform: uppercase; }}
  .stat-val.green  {{ color: #22c55e; }} .stat-val.blue {{ color: #3b82f6; }}
  .stat-val.orange {{ color: #f97316; }} .stat-val.gray {{ color: #94a3b8; }}
  .table-wrap {{ background: white; border-radius: 12px;
                 box-shadow: 0 1px 3px rgba(0,0,0,0.08); overflow: auto; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 12.5px; }}
  th {{ background: #065f46; color: white; padding: 10px; font-weight: 700;
        font-size: 11px; text-transform: uppercase; letter-spacing: 0.04em;
        white-space: nowrap; text-align: left; position: sticky; top: 0; }}
  td {{ padding: 9px 10px; border-bottom: 1px solid #f1f5f9; white-space: nowrap; }}
  tr:hover td {{ background: #f0fdf4; }}
  tr:last-child td {{ border-bottom: none; }}
  .legend {{ margin-top: 16px; font-size: 11px; color: #64748b; line-height: 1.8; }}
</style>
</head>
<body>
<div class="header">
  <h1>🚀 Recent Doublers Scanner</h1>
  <p>{run_date} — {len(results)} titoli raddoppiati — Universo S&amp;P 1500</p>
</div>
<div class="method-note">
  <b>Filtri:</b> Raddoppio ≥ 100% in almeno un timeframe (3/6/9/12 mesi) &nbsp;·&nbsp;
  Distanza dal 52w High ≤ 30% (non collassato) &nbsp;·&nbsp;
  RS Rating ≥ 70/99 (leader) &nbsp;·&nbsp;
  Sopra SMA50 &nbsp;·&nbsp;
  Volume medio ≥ 200k
</div>
{stats_html}
<div class="table-wrap">
<table>
<thead>
<tr>
  <th>Grade</th><th>Score</th>
  <th title="Timeframe più veloce in cui ha raddoppiato">Velocità</th>
  <th title="RS Rating 12m — percentile 1-99 (IBD-style)">RS</th>
  <th>Ticker</th><th>Settore</th>
  <th title="Rendimento negli ultimi 3 mesi">Ret 3m</th>
  <th title="Rendimento negli ultimi 6 mesi">Ret 6m</th>
  <th title="Rendimento negli ultimi 12 mesi">Ret 12m</th>
  <th title="Miglior rendimento nel timeframe più favorevole">Best Ret</th>
  <th title="Distanza dal massimo a 52 settimane">Dist 52wH</th>
  <th>Prezzo</th>
  <th title="Sopra SMA50 — ancora in uptrend">SMA50</th>
  <th>Vol Medio</th><th>Mkt Cap</th>
</tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>
</div>
<div class="legend">
  <b>Grade:</b> A+ = raddoppio in ≤3 mesi | A = ≤6 mesi | B+ = ≤9 mesi | B = ≤12 mesi &nbsp;|&nbsp;
  <b>Best Ret:</b> rendimento massimo nel timeframe migliore &nbsp;|&nbsp;
  <b>Dist 52wH:</b> quant'è distante dal massimo a 52 settimane (verde = vicino, rosso = lontano)
</div>
</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")
    log.info(f"HTML salvato: {output_path}")


# ── Excel ──────────────────────────────────────────────────────────────────────

def generate_recent_doublers_excel(results: list[dict], run_date: str,
                                    output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Doublers {run_date}"

    hdr_fill  = PatternFill("solid", fgColor="065F46")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="E2E8F0")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    grade_fills = {
        "A+": PatternFill("solid", fgColor="86EFAC"),
        "A":  PatternFill("solid", fgColor="93C5FD"),
        "B+": PatternFill("solid", fgColor="FDBA74"),
        "B":  PatternFill("solid", fgColor="D1D5DB"),
    }
    grade_fonts = {
        "A+": Font(bold=True, color="14532D"),
        "A":  Font(bold=True, color="1E3A5F"),
        "B+": Font(bold=True, color="7C2D12"),
        "B":  Font(bold=True, color="374151"),
    }
    green_font = Font(color="15803D", bold=True)
    red_font   = Font(color="DC2626", bold=True)
    gray_font  = Font(color="64748B")
    alt_fill   = PatternFill("solid", fgColor="F0FDF4")

    headers = [
        "Grade", "Score", "Velocità", "RS Rating",
        "Ticker", "Azienda", "Settore",
        "Ret 3m %", "Ret 6m %", "Ret 12m %", "Best Ret %",
        "Dist 52wH %", "52w High ($)", "Prezzo ($)", "SMA50 ($)", "Sopra SMA50",
        "Vol Medio 50gg", "Mkt Cap (B$)",
        "TradingView",
    ]
    col_widths = [
        9, 9, 10, 10,
        9, 26, 22,
        10, 10, 11, 11,
        12, 13, 11, 11, 12,
        16, 13,
        20,
    ]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for ri, r in enumerate(results, 2):
        tv_url = f"https://www.tradingview.com/chart/?symbol={r['ticker']}"
        mc_b   = round(r["market_cap_m"] / 1000, 2) if r["market_cap_m"] else 0.0
        is_alt = (ri % 2 == 0)

        row_vals = [
            r["quality"], r["doubler_score"], r["fastest_tf"], r["rs_rating"],
            r["ticker"], r["name"], r["sector"],
            r["ret_3m"] / 100 if r["ret_3m"] else None,
            r["ret_6m"] / 100 if r["ret_6m"] else None,
            r["ret_12m"] / 100 if r["ret_12m"] else None,
            r["best_return"] / 100,
            r["dist_52w_pct"] / 100,
            r["high_52w"], r["price"], r["sma50"],
            "✓" if r["above_sma50"] else "✗",
            r["avg_vol_50"], mc_b, tv_url,
        ]

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bdr
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_alt: cell.fill = alt_fill

        # Grade
        ws.cell(row=ri, column=1).fill = grade_fills.get(r["quality"], PatternFill())
        ws.cell(row=ri, column=1).font = grade_fonts.get(r["quality"], Font(bold=True))

        # RS Rating
        rs = r["rs_rating"]
        ws.cell(row=ri, column=4).fill = (
            PatternFill("solid", fgColor="86EFAC") if rs >= 90 else
            PatternFill("solid", fgColor="93C5FD") if rs >= 80 else
            PatternFill("solid", fgColor="FDE68A") if rs >= 70 else PatternFill())
        ws.cell(row=ri, column=4).font = Font(bold=True)

        # Ritorni — formato %
        for ci, col in enumerate([8, 9, 10, 11, 12], 8):
            cell = ws.cell(row=ri, column=col)
            cell.number_format = "0.0%"
            if col in [8, 9, 10] and cell.value and float(cell.value or 0) >= 1.0:
                cell.font = green_font
            elif col == 11:
                cell.font = green_font

        # Prezzi
        for col in [13, 14, 15]:
            ws.cell(row=ri, column=col).number_format = "$#,##0.00"

        # Volume e Mkt Cap
        ws.cell(row=ri, column=17).number_format = "#,##0"
        ws.cell(row=ri, column=18).number_format = "#,##0.00"

        # TradingView link
        lc = ws.cell(row=ri, column=19)
        lc.hyperlink = tv_url
        lc.font = Font(color="3B82F6", underline="single")

        ws.row_dimensions[ri].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results)+1}"
    wb.save(str(output_path))
    log.info(f"Excel salvato: {output_path}")
